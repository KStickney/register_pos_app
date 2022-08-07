#Excel Sheet Database
import pandas
from openpyxl import load_workbook
from datetime import datetime, timedelta
import datetime as dt
from PyQt5.QtWidgets import *
from PyQt5 import QtCore
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
from threading import Timer
from shutil import copy2
#from PyQt5.QtGui import QPainter, QPen

# TODO: copy excel file once pass certain date
#TODO: create error log

#Store open new row in a cell?
#PRICE - NEED TO TOTAL ALL FOR THE DAY - save rows indexes that begin and end for that day

new_row = 0 ###ONLY NEEDED FOR FIRST TIME -
date_format = '%d-%m-%Y'
date_format_qt = 'MM dd yyyy'
file = "Pier 142 Database 2.xlsx"
file2 = "Pier 142 Database 2 Backup.xlsx"
wb = load_workbook(file)



cdate       = 'A'
cname       = 'B'
cdepartment = 'C'
cflavor     = 'D'
cextras     = 'E'
cnotes      = 'F'
cquantity   = 'G'
cprice      = 'H'
cfulfilled  = 'I'

crow = 'J'
cnum = 'K'
sheet_number = int(wb["Main"][cnum+"2"].value)
if sheet_number == 0:
    sheet = wb["Main"]
else:
    sheet = wb["Main"+str(sheet_number)]

start_row = 3
cnew_row = 'J2'
MAX_ROWS = 1048575

department_menu = []
flavor_menu = []
extras_menu = []
label_menu = []

def create_worksheet():
    global sheet_number,sheet, wb
    sheet_number += 1
    wb["Main"][cnum + "2"] = sheet_number
    wb.create_sheet("Main"+str(sheet_number),0)
    sheet = wb["Main"+str(sheet_number)]

    sheet[cdate+"1"] = "Date"
    sheet[cname+"1"] = "Name"
    sheet[cdepartment+"1"] = "Department"
    sheet[cflavor+"1"] = "Flavor"
    sheet[cextras+"1"] = "Extras"
    sheet[cnotes+"1"] = "Notes"
    sheet[cquantity+"1"] = "Quantity"
    sheet[cprice+"1"] = "Price"
    sheet[cfulfilled+"1"] = "Fulfilled"
    sheet[crow+"1"] = "Row Index"
    sheet[cnew_row] = start_row

    #globals().update(locals())

def workbook_init():
    pass

    globals().update(locals())

def get_menu():
    sheet = wb["Menu"]
    START_ROW = 2
    row = START_ROW
    while True:
        index = row-2

        #Department
        #department_menu.append([])
        department = sheet['A'+str(row)].value
        if department == '' or department is None:
            break #means no more
        text = ''
        price = 0
        for let in department:
            try:
                price += int(let)
            except:
                if let == ',':
                    pass
                else:
                    text += let
        department_menu.append([text,price])

        #Start of flavor and extras addition
        label_menu.append([])
        flavor_menu.append([])
        extras_menu.append([])

        flavor = sheet['B'+str(row)].value
        flavor = flavor.replace(", ",',')

        flavor_label,flavors = read_extra_cell(flavor)
        label_menu[index].append(flavor_label)

        i=0
        while i < len(flavors):
            try:
                flavor_menu[index].append([flavors[i],int(flavors[i+1])])
                i+=2
            except:
                flavor_menu[index].append([flavors[i], 0])
                i+=1
        #for i in range(0,len(flavors),2):#this assumes given price for each
            #flavor_menu[index].append([flavors[i],int(flavors[i+1])])

        START_COL = 67
        col = START_COL
        k=0
        while True:
            extra = sheet[chr(col) + str(row)].value
            if extra == '' or extra is None:
                break
            extra = extra.replace(", ", ',')
            extras_menu[index].append([])
            label,extras = read_extra_cell(extra)
            label_menu[index].append(label)
            i = 0
            while i < len(extras):
                try:
                    extras_menu[index][k].append([extras[i], int(extras[i + 1])])
                    i+=2
                except:
                    extras_menu[index][k].append([extras[i], 0])
                    i+=1
            #for i in range(0,len(extras),2):
                #extras_menu[index][k].append([extras[i],int(extras[i+1])])
            col+=1
            k+=1

        row += 1
def read_extra_cell(string):
    label = ''
    for let in string:
        if let == ':':
            label+=let
            break
        else:
            label += let
    string = string.replace(label + " ", '')
    string = string.replace(label, '')

    text = ''
    words = []
    for let in string:
        if let == ',':
            if text != '':
                words.append(text)
                text = ''
        else:
            text+=let
    if text != '':
        words.append(text)
    return label,words


def check_old_orders():
    #checks for any unfulfilled orders
    i = start_row
    not_fulfilled_orders = []
    while True:
        if sheet[cdate+str(i)].value is not None and sheet[cfulfilled+str(i)].value is None:
            not_fulfilled_orders.append(i)
        elif sheet[cdate+str(i)].value is None:
            new_row = i
            sheet[cnew_row] = i
            break
        i += 1
    wb.save(file)

def submit_order_data(name,cart):
    try:
        for order in cart:
            new_row1 = new_row
            new_row2 = int(sheet[cnew_row].value)
            if new_row1 > new_row2:
                n_row = new_row1
            else:
                n_row = new_row2

            if n_row > MAX_ROWS:
                create_worksheet()
                n_row = sheet[cnew_row].value

            n_row = str(n_row)



            flavor = ""
            for n in order[1]:
                if flavor == "":
                    flavor += n
                else:
                    flavor += " " + n
        
            extras = ""
            for e in order[2]:
                if extras == "":
                    extras += e
                else:
                    extras += " "+e

            price = 0
            for p in order[5]:
                if type(p)==list:
                    for x in p:
                        price += x
                else:
                    price += p
            price *= order[4]

            sheet[cdate+n_row]       = datetime.today().strftime(date_format)
            sheet[cname+n_row]       = name
            sheet[cdepartment+n_row] = order[0]
            sheet[cflavor+n_row]     = flavor
            sheet[cextras+n_row]     = extras
            sheet[cnotes+n_row]      = order[3]
            sheet[cquantity+n_row]   = order[4]
            sheet[cprice+n_row]      = price

            sheet[cnew_row] = int(n_row) + 1
            
            wb.save(file)

    except Exception as e:
        print(e)

def backorder():
    #put order into excel - but too many orders already on screen
    pass

###Write method how put data for specific range of dates, or a specific week, month, year, etc - do dropdown lists with day,month,year and another for month or year or week
class Sales(QWidget):
    def __init__(self):
        super().__init__()

        #self.setWindowTitle("Pier 142 Sales")

        #self.screen = app.primaryScreen()
        #self.swidth = self.screen.size().width()
        #self.sheight = self.screen.size().height()

        self.styleSheet = f"""
        
        QPushButton#date-submit{{
            background-color: rgba(205,205,205,1);
            border: 1px solid black;
            font: bold 12px;
            border-radius: 12px;
            max-height: 25px;
            min-height: 25px;
            max-width: 100px;
            min-width: 100px;
        }}
        QPushButton#date-submit:pressed{{
            background-color: grey;
            color: white;
        }}
        QPushButton#date-btn{{
            background-color: rgba(205,205,205,1);
            border: 1px solid black;
            font: bold 12px;
            border-radius: 12px;
            max-height: 25px;
            min-height: 25px;
            max-width: 100px;
            min-width: 100px;
        }}
        QPushButton#date-btn:pressed{{
            background-color: grey;
            color: white;
        }}

        QLabel{{
            font: bold 12px;
            max-width: 100px;
            min-width: 50px;
        }}
        
        QLabel#date-label{{
            font: bold 12px;
            max-width: 250px;
            min-width: 250px;
        }}

        QDateEdit{{
            max-width: 100px;
            min-width: 100px;        
        }}
        """

        self.setStyleSheet(self.styleSheet)

        self.UIComponents()
        self.showMaximized()

    def UIComponents(self,):

        self.start_date_edit = QDateEdit(calendarPopup = True)
        self.end_date_edit = QDateEdit(calendarPopup = True)
        self.start_date_edit.setDisplayFormat(date_format_qt)
        self.end_date_edit.setDisplayFormat(date_format_qt)
        self.start_date_edit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.end_date_edit.setDateTime(QtCore.QDateTime.currentDateTime())

        start_date_label = QLabel("Start Date:")
        start_date_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        end_date_label = QLabel("End Date:")
        end_date_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

        submit_button = QPushButton("Submit")
        submit_button.clicked.connect(self.refresh)
        submit_button.setObjectName('date-submit')

        self.date_range_label = QLabel("")
        self.date_range_label.setAlignment(QtCore.Qt.AlignCenter)
        self.date_range_label.setObjectName('date-label')

        self.date_edit_layout = QHBoxLayout()
        self.date_edit_layout.addWidget(start_date_label,alignment=QtCore.Qt.AlignRight)
        self.date_edit_layout.addWidget(self.start_date_edit)
        self.date_edit_layout.addWidget(end_date_label,alignment=QtCore.Qt.AlignRight)
        self.date_edit_layout.addWidget(self.end_date_edit)
        self.date_edit_layout.addWidget(submit_button)
        self.date_edit_layout.addWidget(self.date_range_label)
        self.date_edit_layout.setAlignment(QtCore.Qt.AlignLeft)

        dates = ['Today','Yesterday','Week','Month','Year']
        self.date_btns = []
        date_btn_layout = QHBoxLayout()
        for date in dates:
            btn = QPushButton(date)
            date_btn_layout.addWidget(btn)
            btn.clicked.connect(self.change_date)
            btn.setObjectName('date-btn')
            self.date_btns.append(btn)
        date_btn_layout.setAlignment(QtCore.Qt.AlignLeft)

        date_layout = QVBoxLayout()
        date_layout.addLayout(self.date_edit_layout)
        date_layout.addLayout(date_btn_layout)

        self.tab_widget = QTabWidget()
        self.table_tab = QWidget()
        self.chart_tab = QWidget()
        self.tab_widget.addTab(self.table_tab,"Table")
        self.tab_widget.addTab(self.chart_tab,"Graph")

        self.main_layout = QVBoxLayout()
        self.main_layout.addLayout(date_layout)
        self.main_layout.addWidget(self.tab_widget)


        self.setLayout(self.main_layout)

        self.refresh()

    def change_date(self):
        try:
            btn = self.sender()
            end_date = datetime.today()
            if btn == self.date_btns[0]:
                start_date = end_date
            elif btn == self.date_btns[1]:
                start_date = end_date - timedelta(days=1)
                end_date = end_date - timedelta(days = 1)
            elif btn == self.date_btns[2]:
                start_date = end_date - timedelta(days=end_date.weekday()+1)
            elif btn == self.date_btns[3]:
                start_date = end_date.replace(day=1)
            elif btn == self.date_btns[4]:
                start_date = end_date.replace(month=1,day=1)

            self.start_date_edit.setDateTime(start_date)
            self.end_date_edit.setDateTime(end_date)
        except Exception as e:
            print(e)

    def refresh(self):
        try:
            global wb
            start_date = self.start_date_edit.date().toPyDate().strftime("%m-%d-%Y")
            end_date = self.end_date_edit.date().toPyDate().strftime("%m-%d-%Y")
            self.date_range_label.setText("Date Range: "+start_date+" - "+end_date)
            start_date = datetime.strptime(start_date, "%m-%d-%Y")
            end_date = datetime.strptime(end_date, "%m-%d-%Y")

            sheet1 = wb["Main"]
            for sh in range(0,sheet_number+1):
                if sh == 0:
                    sheet1 = wb["Main"]
                else:
                    sheet1 = wb["Main"+str(sh)]

                start_row_i = int(sheet1[cnew_row].value) - 1
                end_row = start_row_i

                while True:
                    if start_row_i == start_row:
                        break
                    if datetime.strptime(sheet1[cdate + str(end_row)].value, date_format) > end_date:
                        end_row -= 1
                    if datetime.strptime(sheet1[cdate + str(start_row_i)].value, date_format) >= start_date:
                        start_row_i -= 1
                    else:
                        break

                skip = []
                for i in range(1,start_row_i):
                    skip.append(i)
                for i in range(end_row,int(sheet1[cnew_row].value)-1):
                    skip.append(i)

                if sh == 0:
                    self.df = pandas.read_excel(file,sheet_name = "Main",header = 0,skiprows=skip, usecols=cdate+":"+cfulfilled)
                else:
                    df = pandas.read_excel(file,sheet_name = "Main"+str(sh),header = 0,skiprows=skip, usecols=cdate+":"+cfulfilled)
                    self.df = self.df.append(df,ignore_index=True)


            index = self.tab_widget.currentIndex()
            self.tab_widget.removeTab(0)
            self.tab_widget.insertTab(0, Table(self.df), "Table")
            self.tab_widget.removeTab(1) #TODO: Remove once get PyQtGraph
            self.tab_widget.insertTab(1, Chart(self.df),"Graph")
            self.tab_widget.setCurrentIndex(index)
        except Exception as e:
            print(e)

        #print((df['Department']=='Smoothie').sum())
        #print(df.isin(['Chocolate ']).sum(axis=0))
        #for value in df["Department"]:
            #d = "Milk" in str(value)
        #f = df.value_counts("Department")
        #f.keys()

    def get_data(date_range):
        get_row = int(sheet[cnew_row])-1

        if date_range.lowercase() == "year":
            date_days = 365
        elif date_range.lowercase() == 'month':
            date_days = 30
        elif date_range.lowercase() == 'week':
            date_days = 7
        elif date_range.lowercase() == 'yesterday':
            date_days = 1
        elif date_range.lowercase() == 'today':
            date_days = 0

        lower_date = datetime.today() - datetime.timedelta(days = date_days)
        lower_date = lower_date.strftime(date_format)

        #get data by comparing date of each row with lower_date range
        raw_data = []
        while True:
            if datetime.strptime(sheet[cdate+str(get_row)].value,date_format) >= lower_date:
                raw_data.append([sheet[cdepartment+str(get_row)],
                                      sheet[cflavor+str(get_row)],
                                      sheet[cextras+str(get_row)],
                                      sheet[cquantity+str(get_row)],
                                      sheet[cprice+str(get_row)]
                                      ])
            else:
                break
            get_row -= 1

        return raw_data 

    def analyze_data(raw_data):
        #PASS list or just use self.raw_data??
        try:
            uncooked = [[0]]
            for item in raw_data:
                if item[0] not in uncooked:
                    uncooked.append([item[0]])
                #for article in item[0]:

        except Exception as e:
            print(e)


            
class Table(QWidget):
    def __init__(self,df):
        super().__init__()
        df=df.fillna('')
        self.UIComponents(df)

    def UIComponents(self,df):
        try:
            self.table = QTableWidget()
            self.table.setColumnCount(len(df.keys()))
            self.table.setRowCount(len(df[df.keys()[0]])+1)

            i = 0
            for col in df:
                #self.table.insertColumn(i)
                self.table.setItem(0, i, QTableWidgetItem(str(df.keys()[i])))
                for r in range(len(df[col])):
                    self.table.setItem(r+1,i,QTableWidgetItem(str(df[col][r])))
                i+=1

            header = self.table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            #header.setSectionResizeMode(1, QHeaderView.Stretch)

            layout = QVBoxLayout()
            layout.addWidget(self.table)
            self.setLayout(layout)
        except Exception as e:
            print(e)



class Chart(QWidget):
    def __init__(self,df):
        try:
            super().__init__()

            self.styleSheet = f'''
                QLabel#no-data{{
                    font: bold 14px;
                    min-width: 600px;
                }}
            '''
            self.setStyleSheet(self.styleSheet)
            self.UIComponents(df)
        except Exception as e:
            print(e)

    def UIComponents(self,df):
        try:
            if df.empty:
                label = QLabel("There are no orders for the time frame selected")
                label.setObjectName('no-data')
                label.setAlignment(QtCore.Qt.AlignCenter)
                layout = QVBoxLayout()
                layout.addWidget(label)
            else:
                dep = df.value_counts("Department")
                money = df["Price"]
                total_price = 0
                for mon in money:
                    total_price += int(mon)

                data_grid_layout = QGridLayout()
                for row in range(len(dep.keys())):
                    data_grid_layout.addWidget(QLabel(str(dep.keys()[row])),row,0)
                    data_grid_layout.addWidget(QLabel(str(dep[row])), row, 1)
                data_grid_layout.addWidget(QLabel("Total"),row+1,0)
                data_grid_layout.addWidget(QLabel("$"+str(total_price)), row + 1, 1)


                layout = QVBoxLayout()
                layout.addWidget(self.create_piechart(dep))
                layout.addLayout(data_grid_layout)
            layout.setAlignment(QtCore.Qt.AlignCenter)
            self.setLayout(layout)
        except Exception as e:
            print(e)

    def create_piechart(self,dep):
        try:
            series = QPieSeries()
            for k,value in zip(dep.keys(),dep):
                series.append(k,int(value))
            series.setLabelsVisible(True)
            series.setLabelsPosition((QPieSlice.LabelInsideHorizontal))
            series.setLabelsPosition((QPieSlice.LabelOutside))# LabelInsideHorizontal))
            #for slice in series.slices():
                #slice.setLabel("{:.2f}%".format(100 * slice.percentage()))

            piechart = QChart()
            piechart.addSeries(series)
            piechart.setAnimationOptions(QChart.SeriesAnimations)
            piechart.setTitle("Items Sold")
            #piechart.legend().markers(series)[0].setLabel("Python")
                #For Setting Legend names


            piechartview = QChartView(piechart)
            #chartview.setRenderHint(QPainter.Antialiasing)

            return piechartview


        except Exception as e:
            print(e)

def copy_excel():
    copy2(file,file2)

#start_time = Timer(604800,copy_excel) #TODO: Undo before final edition
#start_time.start()
#copy_excel()